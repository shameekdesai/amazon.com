# import openpyxl module 
import openpyxl 

class Excel:
    def __init__(self):
        self.wb = openpyxl.load_workbook("c123.xlsx") 
        self.sheet = self.wb.active   # Get workbook active sheet from active attribute 
        self.r = self.sheet.max_row
        self.c = self.sheet.max_column

    # to read the link from the column
    def readlink(self):
        for i in range(2,self.r):
            d = self.sheet.cell(row = i, column = 4)
            if d.value is None:
                print("no links found at row %d" % i)
            else:
                return d.value
    

    #to write the value in the  coloumn 
    def writetocolumn(self):
        for j in range(2,self.r):
            f = self.sheet.cell(row = j, column = self.c-1)
            #if f.value is None:
            f.value = "5000"
        
        # the save() workbook method. 
        self.wb.save("c123.xlsx") 


#exe = Excel()
#exe.readlink()
